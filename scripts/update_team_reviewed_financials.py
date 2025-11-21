#!/usr/bin/env python
"""
Update financial classification and monetary values in the team-reviewed Excel file
based on the latest merged export (reports/AG_Verification_Summary_FIXED4.xlsx).

Usage:
    python scripts/update_team_reviewed_financials.py
"""

import sys
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
LATEST_EXPORT = PROJECT_ROOT / "reports" / "AG_Verification_Summary_FIXED6.merged.xlsx"
TEAM_REVIEWED_INPUT = PROJECT_ROOT / "reports" / "team_reviewed" / "2025-11-11 AG_Verification_Deduped_Reviewed v0_01NG.xlsx"
TEAM_REVIEWED_OUTPUT = PROJECT_ROOT / "reports" / "team_reviewed" / "AG_Verification_Deduped_Reviewed_and_fixed.xlsx"

SHEET_NAME = "Snippet Raw Data"


def normalize_company(value: str | None) -> str:
    return (value or "").strip().upper()


def load_mapping():
    if not LATEST_EXPORT.exists():
        raise FileNotFoundError(f"Latest export not found: {LATEST_EXPORT}")

    wb = load_workbook(LATEST_EXPORT, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {LATEST_EXPORT}")
    ws = wb[SHEET_NAME]

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_idx = {name: idx for idx, name in enumerate(header)}

    required_cols = ["Company", "Snippet ID", "Financial Type", "Financial Amounts"]
    missing = [col for col in required_cols if col not in header_idx]
    if missing:
        raise ValueError(f"Missing columns in export: {missing}")

    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        company = normalize_company(row[header_idx["Company"]])
        snippet_id = (row[header_idx["Snippet ID"]] or "").strip()
        if not company or not snippet_id:
            continue

        financial_type = row[header_idx["Financial Type"]] or ""
        financial_amounts = row[header_idx["Financial Amounts"]] or ""
        mapping[(company, snippet_id)] = {
            "financial_type": financial_type,
            "financial_amounts": financial_amounts,
        }

    wb.close()
    return mapping


def update_team_reviewed(mapping):
    if not TEAM_REVIEWED_INPUT.exists():
        raise FileNotFoundError(f"Team reviewed file not found: {TEAM_REVIEWED_INPUT}")

    wb = load_workbook(TEAM_REVIEWED_INPUT)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {TEAM_REVIEWED_INPUT}")

    ws = wb[SHEET_NAME]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_idx = {name: idx for idx, name in enumerate(header)}

    required_cols = ["Company", "Snippet ID", "Financial Type", "Financial Amounts"]
    missing = [col for col in required_cols if col not in header_idx]
    if missing:
        raise ValueError(f"Missing columns in team reviewed file: {missing}")

    updated = 0
    missing = 0

    for row in ws.iter_rows(min_row=2):
        company = normalize_company(row[header_idx["Company"]].value)
        snippet_id_cell = row[header_idx["Snippet ID"]]
        snippet_id = (snippet_id_cell.value or "").strip()

        if not company or not snippet_id:
            continue

        key = (company, snippet_id)
        if key not in mapping:
            missing += 1
            continue

        data = mapping[key]
        ft_cell = row[header_idx["Financial Type"]]
        amounts_cell = row[header_idx["Financial Amounts"]]

        if ft_cell.value != data["financial_type"] or (amounts_cell.value or "") != data["financial_amounts"]:
            ft_cell.value = data["financial_type"]
            amounts_cell.value = data["financial_amounts"]
            updated += 1

    output_path = TEAM_REVIEWED_OUTPUT
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    return updated, missing


def main():
    try:
        mapping = load_mapping()
        updated, missing = update_team_reviewed(mapping)
        print(f"[SUCCESS] Updated rows: {updated}")
        if missing:
            print(f"[INFO] Rows without matching data: {missing}")
        print(f"[INFO] Fixed workbook saved to: {TEAM_REVIEWED_OUTPUT}")
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
